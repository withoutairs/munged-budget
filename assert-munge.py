import pandas as pd
import unittest
import yaml
import glob
import datetime
import calendar
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class Munger():
    "Adds our budget category and generally cleans up the Mint data"
    def __init__(self, filename=None, since=2022):
        fname = sorted(glob.glob('*.csv'))[-1]
        csv = fname
        print("Reading from %s, all transactions since %s" % (csv, str(since)))
        self.df = pd.read_csv(csv)
        self.df['date'] = pd.to_datetime(self.df['date'], errors='coerce')
        self.df['year'] = self.df['date'].apply(lambda x: x.year)
        self.df['category'] = self.df['category.name'].apply(lambda x: str(x).lower())
        self.df['month'] = self.df['date'].apply(lambda x: x.month)
        self.df['month-year'] = self.df['date'].apply(lambda x: str(x.year) + '-' + ("%02d" % (x.month)))
        self.df['year-week'] = self.df['date'].apply(lambda x: str(x.year) + '-' + ("%02d" % x.weekofyear))
        self.df = self.df[self.df.year >= since]

        self.file_day = int(fname[26:28])
        self.file_month = int(fname[23:25])
        self.file_year = int(fname[18:22])

    def fix_paychecks(self, row):
        if row['isExpense'] == 'True':
            return row['amount'] * -1
        else:
            return row['amount']

    def munge(self):
        file = open('category_map.yaml', 'r')
        map = yaml.load(file, Loader=yaml.FullLoader)
        file.close()
        self.df['budget_category'] = self.df['category'].apply(map.get)
        self.df['amount'] = self.df.apply(lambda row: self.fix_paychecks(row), axis=1)
        self.df['description'] = self.df.apply(lambda row: self.fix_descriptions(row), axis=1)
        wb = load_workbook('spending-master.xlsx')
        ws = wb.create_sheet(title="Data")
        for r in dataframe_to_rows(self.df, index=False, header=True):
            ws.append(r)
        wb.save('spending.xlsx')

    def fix_descriptions(self, row):
        if row['description'].startswith('OCULUS *'):
            return "Oculus"
        if row['description'].startswith('RALLY HEALTH INC '):
            return "RALLY HEALTH INC "
        if row['description'].startswith('Prime Video'):
            return "Prime Video"
        if row['description'].startswith('POTOMAC ELEC'):
            return "POTOMAC ELEC"
        if row['description'].startswith('To Loan 01'):
            return "To Loan 01"
        if row['description'].startswith('AMZNGrocery'):
            return "AMZNGrocery"
        if row['description'].startswith('UNITED 016'):
            return "UNITED"

        if row['description'].startswith('Prime Now'):
            return "Prime Now"
        if row['description'].startswith('PrimeNowTips'):
            return "Prime Now"
        if row['description'].startswith('PrimeNowMktp'):
            return "Prime Now"

        else:
            return row['description']

class AssertMungeTestCase(unittest.TestCase):
    @classmethod
    def setUpClass(self):
        self.m = Munger()
        self.m.munge()

    def setUp(self):
        with open('budget.yaml', 'r') as f:
            self.budget = yaml.load(f, Loader=yaml.FullLoader)

    def test_all_transactions_have_budget_categories(self):
        no_budget_category = self.m.df[self.m.df.budget_category.isnull()]
        pd.set_option('display.max_columns', 500)
        no_budget_category = no_budget_category[['date', 'description', 'amount', 'category']]
        if len(no_budget_category) > 0:
            print("The following transactions have no budget category")
            print(no_budget_category)
            unittest.TestCase.assertEqual(self, 0, len(no_budget_category))

    def test_not_too_many_budget_categories(self):
        grouped = self.m.df.groupby(['budget_category'])
        too_many = 30
        how_many = len(grouped.aggregate(sum))
        unittest.TestCase.assertLessEqual(self, how_many, too_many,
                                          "How can you think about %s budget categories?" % how_many)

    def get_budget_goal(self, row):
        goal = self.budget.get(row['budget_category'])
        if goal is not None:
            return goal
        else:
            print("**** no budget goal for " + (row['budget_category']))

    def format_budget_progress(self, row):
        return str(round((row['amount'] / row['budget_goal']) * 100)) + '%'

    def is_over(self, row, p):
        if p <= (row['amount'] / row['budget_goal']):
            return "*"
        else:
            return ""

    def test_against_budget(self):
        print('x' * 80)
        p = int(self.m.file_day) / calendar._monthlen(self.m.file_year, self.m.file_month)
        p_fmt = str(int(round(p, 2) * 100))
        print(f"File is {p_fmt}% of the month")

        # month_wanted = datetime.datetime.now().isoformat()[:7]  # usually
        month_wanted = '2022-05' # override
        budget_category_totals = self.m.df[self.m.df['month-year'] == str(month_wanted)].groupby('budget_category').sum('amount') * -1
        budget_category_totals['budget_category'] = budget_category_totals.index
        budget_category_totals['budget_goal'] = budget_category_totals.apply(lambda row: self.get_budget_goal(row), axis=1)
        budget_category_totals['budget_progress'] = budget_category_totals.apply(lambda row: self.format_budget_progress(row), axis=1)
        budget_category_totals['is_over'] = budget_category_totals.apply(lambda row: self.is_over(row, p), axis=1)
        print(budget_category_totals.loc[:, ['budget_category','amount', 'budget_goal', "budget_progress", "is_over"]])
        print('-=' * 40)

    @classmethod
    def tearDownClass(self):
        os.system("open spending.xlsx")